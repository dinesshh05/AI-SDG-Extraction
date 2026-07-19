import os
from collections import Counter
from difflib import SequenceMatcher

from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font, Alignment  # type: ignore


SDG_TITLES = {
    1:  "SDG 1: No Poverty",
    2:  "SDG 2: End hunger, achieve food security and improved nutrition and promote sustainable agriculture",
    3:  "SDG 3: Good Health and Well-being",
    4:  "SDG 4: Quality Education",
    5:  "SDG 5: Gender Equality",
    6:  "SDG 6: Clean Water and Sanitation",
    7:  "SDG 7: Affordable and Clean Energy",
    8:  "SDG 8: Decent Work and Economic Growth",
    9:  "SDG 9: Industry, Innovation and Infrastructure",
    10: "SDG 10: Reduced Inequalities",
    11: "SDG 11: Sustainable Cities and Communities",
    12: "SDG 12: Responsible Consumption and Production",
    13: "SDG 13: Climate Action",
    14: "SDG 14: Life Below Water",
    15: "SDG 15: Life on Land",
    16: "SDG 16: Peace, Justice, and Strong Institutions",
    17: "SDG 17: Partnerships for the Goals",
}

HEADER_FONT = Font(bold=True, size=10)
BODY_FONT = Font(bold=False, size=10)
NO_ACTIVITY_TEXT = "No activity fetched"

_JUNK_PHRASES = {
    "no specific initiative mentioned",
    "no initiative mentioned",
    "not specified",
    "none mentioned",
    "no activity mentioned",
}

_DEDUP_SIMILARITY_THRESHOLD = 0.80
_MAX_SDGS_PER_INITIATIVE = 3


def _is_junk_initiative(item):
    """
    Filters out placeholder entries the LLM occasionally produces when
    a chunk mentions a topic without describing a concrete action, but
    the model still outputs a row instead of skipping it.
    """
    name = (item.get("initiative_name") or "").strip().lower()
    desc = (item.get("description") or "").strip().lower()
    return name in _JUNK_PHRASES or desc in _JUNK_PHRASES


def _similarity(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def _deduplicate_initiatives(initiatives):
    """
    Deduplicates on initiative_name + description similarity BEFORE
    SDG grouping. Batching and chunk overlap can cause the same
    real-world fact to be independently extracted 2-3 times with
    slightly different wording.

    Each individual LLM call already caps itself at 2-3 SDGs, but if
    several independent extractions of the same fact each picked a
    different (individually valid) SDG subset, a naive union of tags
    across duplicates can push the merged result past that cap. To
    prevent that, SDG tags are tracked by frequency across all merged
    duplicates, and only the most frequently agreed-upon SDGs are
    kept, up to _MAX_SDGS_PER_INITIATIVE - SDGs only one out of
    several independent extractions picked are the weaker signal and
    are dropped first.
    """

    kept = []
    sdg_frequency = []
    sdg_name_lookup = []

    for item in initiatives:

        name = item.get("initiative_name", "")
        desc = item.get("description", "")
        combined_text = f"{name} {desc}"

        matched_index = None

        for i, existing in enumerate(kept):

            existing_text = f"{existing.get('initiative_name', '')} {existing.get('description', '')}"

            if _similarity(combined_text, existing_text) >= _DEDUP_SIMILARITY_THRESHOLD:
                matched_index = i
                break

        item_sdg_ids = item.get("sdg_ids", [])
        item_sdg_names = item.get("sdg_names", [])
        name_map = dict(zip(item_sdg_ids, item_sdg_names))

        if matched_index is not None:

            existing = kept[matched_index]

            sdg_frequency[matched_index].update(item_sdg_ids)
            sdg_name_lookup[matched_index].update(name_map)

            if not existing.get("metric") and item.get("metric"):
                existing["metric"] = item["metric"]

        else:
            kept.append(dict(item))
            sdg_frequency.append(Counter(item_sdg_ids))
            sdg_name_lookup.append(dict(name_map))

    for i, item in enumerate(kept):

        freq = sdg_frequency[i]

        if len(freq) <= _MAX_SDGS_PER_INITIATIVE:
            final_ids = sorted(freq.keys())
        else:
            ranked = sorted(freq.items(), key=lambda x: (-x[1], x[0]))
            final_ids = sorted(sdg_id for sdg_id, _ in ranked[:_MAX_SDGS_PER_INITIATIVE])

        item["sdg_ids"] = final_ids
        item["sdg_names"] = [
            sdg_name_lookup[i].get(sdg_id, "")
            for sdg_id in final_ids
        ]

    return kept


def _group_by_sdg(initiatives):
    """One initiative can map to multiple SDGs, so it's repeated under each."""

    grouped = {sdg_no: [] for sdg_no in SDG_TITLES}

    for item in initiatives:
        for sdg_id in item.get("sdg_ids", []):
            if sdg_id in grouped:
                grouped[sdg_id].append(item)

    return grouped


def _format_activity_list(items):

    if not items:
        return NO_ACTIVITY_TEXT

    lines = []

    for i, item in enumerate(items, start=1):
        text = (item.get("description") or item.get("initiative_name", "")).strip()
        metric = item.get("metric", "")
        if metric:
            text = f"{text} ({metric})" if text else metric
        lines.append(f"{i}. {text}")

    return "\n".join(lines)


def export_to_excel(initiatives, output_path="output/sustainability_report.xlsx"):

    initiatives = [item for item in initiatives if not _is_junk_initiative(item)]
    initiatives = _deduplicate_initiatives(initiatives)

    os.makedirs("output", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sustainability Initiatives"

    grouped = _group_by_sdg(initiatives)
    row = 1

    for sdg_no in sorted(grouped):

        ws.cell(row=row, column=1, value=SDG_TITLES[sdg_no]).font = HEADER_FONT
        row += 1

        cell = ws.cell(row=row, column=1, value=_format_activity_list(grouped[sdg_no]))
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True)
        row += 2

    ws.column_dimensions["A"].width = 100

    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")

    return output_path